#! /usr/bin/env python3

from openpyxl import load_workbook
from astropy.table import Table
import numpy as np
import scipy as sp
import scipy.stats
from matplotlib import pylab as pl
import matplotlib as mpl
import re
import sys

pl.rcParams['text.usetex'] = True
pl.rcParams['font.family'] = 'serif'
pl.rcParams['font.size'] = 10


METRIC = {'U/M': 'x1', 'U/S': 'x2', 'Sp/S': 'x3', 'G/S': 'x4', 'P/S': 'x5'}

def read_all_years():

    tabs = [read(year, fill_missing=True, verbose=True) 
                    for year in range(2006, 2022)]

    # compute funding history

    nuniv = len(tabs[-1])
    AFD_old = np.zeros((nuniv,))
    F = np.zeros((5, nuniv))
    names = ['F1', 'F2', 'F3', 'F4', 'F5']

    for tab in tabs:

        AFD95 = tab['AFD95%']
        dampening = tab['AFD95%'] / np.maximum(AFD_old, 0.1)
        AFD = AFD95 + tab['AFD5%']

        F = dampening * F +  np.array([tab[f"f{i}"] for i in range(1, 6)])

        tab.add_columns(F, names=names)
        tab.add_column(AFD - F.sum(axis=0), name='Fh')

        AFD_old = AFD

    return tabs
    
def read(year, fill_missing=False, verbose=False):

    if verbose:
        print(f'read table for {year}')

    book = load_workbook('../src/tabla-afd.xlsx', read_only=True)
    
    sheet = book[book.sheetnames[1]]
    nuniv = 25 + 2 * (year >= 2018)
    first = 14 + 35 * (2021 - year)
    last = first + nuniv - 1
    area = 'A{}:L{}'.format(first, last)
    values = [list(c.value for c in r) for r in sheet[area]]
    short = ['de', 'del', 'la', 'el', 'las', 'los']
    
    for i, row in enumerate(values):

        uni = row[0]
        uni = re.sub('\.(?=\S)', '. ', uni)
        uni = re.sub('Maria', 'María', uni)
        uni = re.sub('Bio Bio', 'Bío-Bío', uni)
        uni = ' '.join([c.capitalize() if c not in short else c 
                for c in re.split('\s+', uni)])
        uni = re.sub("O'h", "O'H", uni)
        row[0] = uni
        
        for j in range(1, 12):
            row[j] = float(row[j])

    names = ['University', 'U', 'M', 'S', 'Sp', 'G', 'Pi', 'Ps', 'P', 
                '%_AFD5%', 'AFD5%', 'AFD95%']
    tab = Table(rows=values, names=names)
    tab.remove_columns(['Pi', 'Ps'])
  
    # 2010 calculations have been made using 2009 variables, but the table
    # from the Ministry lists 2010 ones.

    if year == 2010:
        tab2 = read(2009)
        for name in ['U', 'M', 'S', 'Sp', 'G', 'P']:
            tab[name] = tab2[name]
    
 
    tab.meta['computed'] = False 
    compute_funding(tab)
    if verbose:
        print(f'   funding computations done')
    
    # to make table of all years the same size

    if fill_missing and year <= 2017:
        nfields = len(tab.colnames) - 1
        tab.add_row(("U. de O'Higgins",) + (0,) * nfields)
        tab.add_row(("U. de Aysén",) + (0,) * nfields)
    
    tab.meta['year'] = year
    tab.meta['computed'] = True 
    
    return tab

def _set_column_group(tab, x, name):
    for k, xk in enumerate(x):
        colname = '{}{}'.format(name, 1 + k)
        if colname in tab.colnames:
            tab[colname] = xk
        else:
            tab.add_column(xk, name=colname)

def _set_column(tab, x, name):
    if name in tab.colnames:
        tab[name] = x
    else:
        tab.add_column(x, name=name)

def compute_funding(tab, skip=None):
   
    if tab.meta['computed']:
        return
 
    c = [0.01, 0.15, 0.24, 0.25, 0.35]
    
    if skip is None or 'x' not in skip:
        x = [tab['U'] / np.maximum(1, tab['M']),
             tab['U'] / tab['S'],
             tab['Sp'] / tab['S'],
             tab['G'] / tab['S'],
             tab['P'] / tab['S']]
        _set_column_group(tab, x, 'x')
    
    else:
    
        x = [tab[n] for n in ['x1','x2','x3','x4','x5']]
    
    if skip is None or 'xi' not in skip:
    
        xi = [(xk - xk.mean())/ xk.std(ddof=0) for xk in x]
        _set_column_group(tab, xi, 'xi')
    
    else:
    
        xi = [tab[n] for n in ['xi1','xi2','xi3','xi4','xi5']]

    y = [np.exp((xik / 4 - 7/5) ** 3) for xik in xi]
    _set_column_group(tab, y, 'y')
    ytot = sum(ck * sum(yk) for ck, yk in zip(c, y))
    
    p = [ck * yk / ytot for ck, yk in zip(y, c)]
    _set_column_group(tab, p, 'p')
    _set_column(tab, sum(pk for pk in p), name='p')
    
    afd5 = np.sum(tab['AFD5%'])
    f = [pk * afd5 for pk in p]
    _set_column_group(tab, f, 'f')
    _set_column(tab, np.round(tab['p'] * afd5), name='f')

    tab.meta['computed'] = True


def change_metrics(tab0, univ, variable, *, increment=1, unit='number'):
    
    tab = tab0.copy()
    tab.meta['computed'] = False

    print(tab.meta['computed'])
    univ = np.atleast_1d(univ)
    variable = np.atleast_1d(variable)

    for univ in univ:
        for variable in variable:
            
            if unit == 'stdev':

                v1, v2 = variable.split('/')
                stdev = tab[METRIC[variable]].std(ddof=0)
                tab[univ][v1] += stdev * increment * tab[univ][v2]

            else:

                tab[univ][variable] += increment

    compute(tab)

    return tab

