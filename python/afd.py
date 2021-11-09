#! /usr/bin/env python3

from matplotlib import pylab as plt
from openpyxl import load_workbook
from astropy.table import Table
import numpy as np
import scipy as sp
import scipy.stats
from matplotlib import pylab as pl
import matplotlib as mpl
import re
import sys

class AFD:

    c = np.array([0.01, 0.15, 0.24, 0.25, 0.35])
    metrics = np.array(['U', 'M', 'S', 'Sp', 'G', 'P'])

    def __init__(self, start = 2006, end = 2020):
        self.years = np.arange(start, end + 1)
        self.nyear = len(self.years)
        self.nuniv = 27
        cols = self.metrics.tolist() + ['AFD5%', 'AFD95%']
        self.nmetrics = len(self.metrics)
        data = np.ma.masked_invalid([read_table(y, compute=False, fill_missing=True)[cols].as_array().tolist() for y in self.years])
        data = np.ma.masked_invalid(data)
        self.afd05 = data[-2].sum(axis=-1)
        self.afd95 = data[-1].sum(axis=-1)
        self._compute()

    def _compute_x(self):
        data = self.data
        self.x = data[...,[0,0,3,4,5]] / np.maximum(1e-3, data[...,[1,2,2,2,2]])

    def _compute_xi(self):
        x = self.x
        mean = x.mean(axis=1)[:,None,:]
        std = x.std(axis=1, ddof=0)[:,None,:]
        self.xi = (x - mean) / std 
        self.xi.fill_value = np.log(1e+20)

    def _compute_y(self):
        self.y = np.exp((self.xi / 4 - 7/5) ** 3) 

    def _compute_p(self):
        ysum = np.sum(self.y * self.c, axis=-1)
        ytot = ysum.sum(axis=-1)
        self.p = y / ytot[:,None,None]
        self.psum = ysum / ytot[:,None]

    def _compute_afd(self):
        this = self.years - self.start
        ref = this - (self.years == 2010) 
        self.data[this,:,6] = self.psum[:,ref] * self.afd05[this,None]
        for i, j in zip(this[1:], ref[:-1]):
            ref_funding = self.afd95[j] + self.afd05[j]
            ref_p = self.data[j,:,6:8].sum(axis=1) / ref_funding
            self.data[i,:,7] = self.afd95[i] * ref_p

    def _compute(self, exclude=None):
        if exclude not in ['x', 'xi', 'y']:
            self._compute_x()
        if exclude not in ['xi', 'y']:
            self._compute_xi()
        if exclude not in ['y']:
            self._compute_y()
        self._compute_p()
        self._compute_afd()

    def change_metric(self, metric, years, values, unit='number'):
        im = np.argwhere(self.metric == metric)[0,0] 
        iy = np.argwhere(years, values)
            

def read_table(year, compute=False, fill_missing=False):
    book = load_workbook('../src/tabla-afd.xlsx', read_only=True)
    sheet = book[book.sheetnames[0]]
    nuniv = 25 + 2 * (year >= 2018)
    first = 15 + 35 * (2020 - year)
    last = first + nuniv - 1
    area = 'A{}:L{}'.format(first, last)
    values = [list(c.value for c in r) for r in sheet[area]]
    short = ['de', 'del', 'la', 'el', 'las', 'los']
    for i, row in enumerate(values):
        uni = row[0]
        uni = re.sub('\.(?=\S)', '. ', uni)
        uni = re.sub('Maria', 'María', uni)
        uni = re.sub('Bio Bio', 'Bío-Bío', uni)
        uni = ' '.join([c.capitalize() if c not in short else c 
                for c in re.split('\s+', uni)])
        uni = re.sub("O'h", "O'H", uni)
        row[0] = uni
        for j in range(1, 12):
            row[j] = float(row[j])
    names = ['University', 'U', 'M', 'S', 'Sp', 'G', 'Pi', 'Ps', 'P', '%_AFD5%', 'AFD5%', 'AFD95%']
    tab = Table(rows=values, names=names)
    tab.remove_columns(['Pi', 'Ps'])
    if compute:
        compute_table(tab)
    if fill_missing and year <= 2017:
        nfields = len(tab.colnames) - 1
        tab.add_row(("U. de O'Higgins",) + (np.nan,) * nfields)
        tab.add_row(("U. de Aysén",) + (np.nan,) * nfields)
    tab.meta['year'] = year
    tab.meta['computed'] = bool(compute)
    return tab

def set_column_group(tab, x, name):
    for k, xk in enumerate(x):
        colname = '{}{}'.format(name, 1 + k)
        if colname in tab.colnames:
            tab[colname] = xk
        else:
            tab.add_column(xk, name=colname)
   
def set_column(tab, x, name):
    if name in tab.colnames:
        tab[name] = x
    else:
        tab.add_column(x, name=name)
 
def compute_table(tab, skip=None):
    c = [0.01, 0.15, 0.24, 0.25, 0.35]
    if skip is None or 'x' not in skip:
        x = [tab['U'] / np.maximum(1, tab['M']),
             tab['U'] / tab['S'],
             tab['Sp'] / tab['S'],
             tab['G'] / tab['S'],
             tab['P'] / tab['S']]
        set_column_group(tab, x, 'x')
    else:
        x = [tab[n] for n in ['x1','x2','x3','x4','x5']]
    if skip is None or 'xi' not in skip:
        xi = [(xk - xk.mean())/ xk.std(ddof=0) for xk in x]
        set_column_group(tab, xi, 'xi')
    else:
        xi = [tab[n] for n in ['xi1','xi2','xi3','xi4','xi5']]
    y = [np.exp((xik / 4 - 7/5) ** 3) for xik in xi]
    set_column_group(tab, y, 'y')
    ytot = sum(ck * sum(yk) for ck, yk in zip(c, y))
    p = [ck * yk / ytot for ck, yk in zip(y, c)]
    set_column_group(tab, p, 'p')
    set_column(tab, sum(pk for pk in p), name='p')
    afd5 = np.sum(tab['AFD5%'])
    f = [pk * afd5 for pk in p]
    set_column_group(tab, f, 'f')
    set_column(tab, np.round(tab['p'] * afd5), name='f')
    tab.meta['computed'] = True

def change_metric(tab, univ, name, increment=1, unit='number'):
    tab = tab.copy()
    univ = np.atleast_1d(univ)
    for univ in univ:
        if unit == 'stdev':
            s = tab[univ]['S']
            if name == 'P':
                tab[univ]['P'] += tab['x5'].std(ddof=0) * increment * s 
            elif name == 'G':
                tab[univ]['G'] += tab['x4'].std(ddof=0) * increment * s
            elif name == 'Sp':
                tab[univ]['Sp'] += tab['x3'].std(ddof=0) * increment * s
            elif name == 'U':
                tab[univ]['U'] += tab['x2'].std(ddof=0) * increment * s
            else:
                raise NotImplementedError('not doable actually')
        else:
            tab[univ][name] += increment
    # recompute table
    compute_table(tab)
    return tab

def marginal_earning(tab, univ, metric, unit=None):
    tab1 = change_metric(tab, univ, metric, unit=unit)
    df = tab['AFD5%'].sum() * (tab1[univ]['p'] - tab[univ]['p'])
    return df

def yearly_marginal_earnings(year, univ, metric, unit=None):
    tab = read_table(year)
    df = [[marginal_earning(tab, u, m, unit=unit) 
                for u in univ]
                    for m in metric]
    f95 = tab['AFD95%'].sum()
    f = f95 + tab['AFD5%'].sum() 
    # in 2018, 95% is not 95% some go to new universities, old ones
    # get a lower share ;-)
    if year == 2018:
        f95 -= tab[[25,26]]['AFD95%'].sum()
    print(year, f95, f)
    return f95, f, df

def historical_marginal_earnings(metric=['G', 'P'], univ=[0,1,2,3], unit=None):
    start = 2006
    end = 2019
    years = np.arange(start, end + 1)
    me = [yearly_marginal_earnings(year, univ, metric, unit=unit)
        for year in years]
    F95, F, dF = [np.array(a) for a in zip(*me)]
    return years, F95, F, dF 

def cumulated_marginal_earnings(y, F95, F, dF, start=2013, duration=3, p=0,
        icorr=0):
    now = max(y)
    start = np.array(start)
    duration = np.array(duration)
    # extrapolation
    y_ex = np.arange(now + 1, (start + duration).max() + 31)
    icorr_ex = np.ones_like(y_ex)
    F_ex = F[-1] * (1 + p) ** (y_ex - now)
    F95_ex = F95[-1] * (1 + p) ** (y_ex - now) 
    dF_ex = extrapolate_earnings(y, dF, y_ex, icorr=icorr)
    # merge past values with extrapolated ones
    dF = np.vstack([dF, dF_ex])
    F = np.hstack([F, F_ex])
    F95 = np.hstack([F95, F95_ex])
    y = np.hstack([y, y_ex])
    icorr = np.hstack([icorr, icorr_ex])
    # recursion
    dF_cum = np.zeros_like(dF)
    for i, cy in enumerate(y):
        # in 2010, 2009 percentages are used, year index j = i - 1.
        j = i - 1 * (cy == 2010)
        # however total fundings changes absolute value linked to % -> f 
        f = F[i] / F[j]
        # this year marginal earnings: is the grant/researcher active?
        active = (cy > start) * (cy <= start + duration)
        dF_cum[i] = dF[j] * f * active[:,None]
        # from 2007, marginal earnings cumulate through the 95%
        if i > 0:
            dF_cum[i] += dF_cum[j - 1] * F95[i] / F[j - 1]
    keep = y >= start.min()
    ic = icorr[:,None,None]
    return y[keep], F95[keep] / icorr, F[keep] / icorr, dF_cum[keep] / ic

#   # univ = [0, 1, 2, 5, 4, 8]
#   univ = np.arange(25)
#   metric = ['Sp', 'G', 'P']
#   nx = 2
#   ny = 3
#   y, F95, F, dF = historical_marginal_earnings(metric=metric, 
#       univ=univ, unit=None)
#   # dF = np.maximum(dF, 1) # one weird case with dF < 1

#    
#   university = read_table(2019)[univ]['University']
#   # university = [re.sub(' ?(Católica|C\.)', 'C.', u) for u in university]
#   university = [re.sub('\.(?=\S)', '. ', u) for u in university]
#   university = [re.sub('Téc\.', 'Técnica', u) for u in university]
#   university = [re.sub('Sta\.', 'Santa', u) for u in university]
#   what = ['prof', 'grant', 'paper'] 
#   uf = Table.read('../src/uf.tsv', format='ascii.csv', delimiter='\t')
#   icorr = uf['UF'][-1] / uf['UF'] 

def extrapolate_earnings(yin, dF, yout, icorr=1):
    ny, nm, nu = dF.shape
    z = np.log(dF * icorr[:,None,None])
    coeff = np.array([[sp.stats.linregress(yin, z[:,m,k])[0:2]
                for m in range(nm)]
                    for k in range(nu)])
    a, b = coeff.T
    for m in range(nm):
        for k in range(nu):
            print(m, k, a[m,k], b[m,k])
    zout = a * yout[:,None,None] + b
    large_future = (zout > zout[yout == 2019]) * (yout[:,None,None] > 2019)
    znow = large_future * zout[yout == 2019]
    f = (zout[large_future] / znow[large_future]) ** -4
    zout[large_future] = ((1 - f) * znow[large_future] + f * zout[large_future])
    return np.exp(zout)

# y0 = np.arange(y[0], y[0] + 30)
# dF_ex = extrapolate_earnings(y, dF, y0, icorr=icorr)
#
#   # keep universities in the first half of prof #
#   for m, me in enumerate(metric):
#       fig = pl.figure(m + 1, figsize=(7.5,10))
#       fig.clf()
#       fig.subplots_adjust(wspace=0, left=0.1, bottom=0.04,
#           hspace=0, top=.99, right=.99)
#       for j in range(ny):
#           for i in range(nx):
#               k = i + nx * j  
#               ax = fig.add_subplot(ny, nx, k + 1)
#               if j < ny - 1:
#                   ax.set_xticklabels([])
#               else:
#                   ax.set_xlabel('year')
#               if i > 0:
#                   ax.set_yticklabels([])
#               else:
#                   ax.set_ylabel('MCLP/{}'.format(what[m]))
#               keep = y != 2010
#               ax.plot(y[keep], dF[keep,m,k] * icorr[keep] / 1e3, 'ko', 
#                       y0, dF_ex[:,m,k] / 1e3, 'k:')
#       ymax = 1.1 * np.max([ax.get_ylim()[1] for ax in fig.axes])
#       for ax, u in zip(fig.axes, university):
#           ax.set_ylim(0, ymax)
#           ax.text(2006, 0.95 * ymax, u, va='top')
#       fig.show()
#       fig.savefig('marginal-earnings-by-{}.pdf'.format(what[m]))
   

# keep universities in the first half of prof #

# p = 0
# y_cum, F95_cum, F_cum, dF_cum = cumulated_marginal_earnings(y, F95, F, dF, 
#     start=[2006, 2016, 2018], duration=[30, 3, 1], p=p, icorr=icorr)
#   for m, me in enumerate(metric):
#       fig = pl.figure(m + 11, figsize=(7.5,10))
#       fig.clf()
#       fig.subplots_adjust(wspace=0, left=0.1, bottom=0.04,
#           hspace=0, top=.99, right=.99)
#       for j in range(ny):
#           for i in range(nx):
#               k = i + nx * j
#               ax = fig.add_subplot(ny, nx, k + 1)
#               if j < ny - 1:
#                   ax.set_xticklabels([])
#               else:
#                   ax.set_xlabel('year')
#               if i > 0:
#                   ax.set_yticklabels([])
#               else:
#                   ax.set_ylabel('MCLP/{}'.format(what[m]))
#               keep = y != 2010
#               ax.plot(y_cum, dF_cum[:,m,k] / 1e3, 'k-'),
#       ymax = 1.1 * np.max([ax.get_ylim()[1] for ax in fig.axes])
#       for k, (ax, u) in enumerate(zip(fig.axes, university)):
#           ax.set_ylim(0, ymax)
#           ax.text(2006, 0.95 * ymax, u, va='top')
#           f = 0.95 * (1 + p)
#           df_tot = (dF_cum[:,m,k].sum() + f/(1-f)*dF_cum[-1,m,k]) / 1e3
#           if me == 'P':
#               txt = 'average = {:.0f}M'.format(df_tot)
#           else:
#               df = df_tot / duration[m] / 12 
#               txt = 'average = {:.1f}M/mo'.format(df)
#           ax.text(y_cum[-2], 0.8 * ymax, txt, ha='right')
#       fig.show()
#       fig.savefig('cumulated-earnings-by-{}.pdf'.format(what[m]))


def cumulated(univ=[[0, 1],[2,3]], start=2006, metric='Sp', ny=30, p=1.00, 
        name=None):
    years = np.arange(start, start + 60)
    fig = pl.figure(1)
    fig.clf()
    maxy = 0
    mean = []
    for iu in range(2):
        for ju in range(2):
            F = []
            dF = []
            dF_tot = []
            k = univ[iu][ju]
            for i, year in enumerate(years):
                if year < 2019:
                    tab = read_table(year)
                    f5 = tab['AFD5%'].sum()
                    f = f5 +  tab['AFD95%'].sum()
                    if year == 2010:
                        tab = read_table(2009)
                    if year < start + ny:
                        tab1 = change_metric(tab, k, metric, increment=1) 
                        df = f5 * (tab1[k]['p'] - tab[k]['p'])
                    else:
                        df = 0
                    df_tot = df
                    if i > 0:
                        df_tot += .95 * dF_tot[-1] * f / F[-1] 
                else:
                    f = F[-1] * p
                    df = np.mean(dF[-4:]) * (year < start + ny)
                    df_tot = .95 * dF_tot[-1] * f / F[-1] + df
                dF.append(df)
                dF_tot.append(df_tot)
                F.append(f)
                #print(year, f, df, df_tot)
            icorr = 1.04 ** np.maximum(0, 2018 - years)
            dF_tot *= icorr
            #print(1 + ju + 2*iu, k)
            ax = fig.add_subplot(2, 2, 1 + ju + 2*iu)
            maxy = np.maximum(1e-3*max(dF_tot), maxy)
            if iu == 0:
                ax.set_xticklabels([])
            if ju == 1:
                ax.set_yticklabels([])
            now = years <= 2019
            ax.plot(years[now], 1e-3*dF_tot[now], 'k-', 
                    years, 1e-3*dF_tot, 'k:')
            if iu == 1:
                ax.set_xlabel('year')
            if ju == 0:
                ax.set_ylabel('2019 million Chilean pesos')
            fig.subplots_adjust(left=0.2, right=0.98, top=0.98, hspace=0, wspace=0)
            fig.show()
            mean.append((sum(dF_tot) + dF_tot[-1] * (20*0.95))/ny/12)
            print(mean)
    for m, ax, k in zip(mean, fig.axes, np.ravel(univ)):
        ax.set_ylim(0, maxy)
        ax.text(years[2], maxy*0.98, tab[k]['University'], va='top')
        ax.text(years[2], maxy*0.02, "avg. {:3.1f} M/mo".format(1e-3*m))
    if name is not None:
        fig.savefig(name)

#cumulated(metric='Sp',start=2006,univ=[[0, 1],[2, 3]],ny=30, name='staff.pdf')
#cumulated(metric='G',start=2016,univ=[[0,1],[2,3]],ny=3, name='postdoc.pdf')

def variation(tabs):
    years = np.array([tab.meta['year'] for tab in tabs])
    names = tabs[0].colnames[1:]
    rows = [(str(tab.meta['year']),) + tuple(float(sum(tab[name])) for name in names) 
                for tab in tabs]
    tab = Table(rows=rows, names=['year', *names])
    tab.remove_column('%_AFD5%')
    tab.add_column(tab['AFD5%'] +  tab['AFD95%'], name='AFD')
    uf = Table.read('../src/macro.tsv', format='ascii.basic')
    growth = uf['growth'][:-1]
    ir_growth = uf['IR'][:-1]
    tab.add_column(uf['UF'], name='UF')
    tab.add_column(tab['AFD'] * (tab['UF'][-1] / tab['UF']), name='AFD_real')
    first, last = tab[0], tab[-1]
    ny = int(last['year']) - int(first['year'])
    increase = [(last[n] / first[n])**(1/ny) - 1 for n in tab.colnames[1:]]
    row = ('increase', *increase)
    tab.add_row(row)
    mean_growth = np.exp(np.mean([np.log(1+g/100) for g in growth])) - 1
    mean_ir_growth = np.exp(np.mean([np.log(1+g/100) for g in ir_growth])) - 1
    f = (1 + growth/100)
    ir_f = (1 + ir_growth/100)
    gdp = (1/f[::-1]).cumprod()[::-1].tolist()
    gdp += [1, mean_growth]
    ir = (1/ir_f[::-1]).cumprod()[::-1].tolist()
    ir += [1, mean_ir_growth]
    tab.add_column(gdp, name='GDP_percapita')
    tab.add_column(ir, name='IR_real')
    fig = pl.figure(1)
    fig.clf()
    ax = fig.add_subplot(111)
    ax2 = ax.twinx()
    grey = (.4,.4,.7)
    line2 = ax2.plot(years, tab['GDP_percapita'][:-1],
            color=grey, linestyle='-.',
            label=f"GDP per cápita ({mean_growth:+.1%})")
    line3 = ax2.plot(years, tab['IR_real'][:-1], 
            color=grey, linestyle=':',
            label=f"mean real wage ({mean_ir_growth:+.1%})")
    line4 = ax2.plot(years-1, tab['U'][:-1] / tab['U'][-2], 
            color=grey, linestyle='--',
            label=f"undergraduates ({tab['U'][-1]:+.1%})")
    line5 = ax2.plot(years-1, tab['S'][:-1] / tab['S'][-2], 
            color=grey, dashes=[5,2,2,2,2,2],
            label=f"professors ({tab['S'][-1]:+.1%})")
    ax2.set_ylim(0, ax2.get_ylim()[1])
    ax2.set_ylabel('relative value')
    ax2.spines['right'].set_color(grey) 
    ax2.tick_params(axis='y', color=grey)
    ax2.yaxis.label.set_color(grey)
    for label in ax2.get_yticklabels():
        label.set_color(grey)
    line1 = ax.plot(years, tab['AFD_real'][:-1] / 1e6, 'k-', 
            label=f"AFD ({tab['AFD_real'][-1]:+.1%})")
    ax.set_ylabel('billions 2020 Chilean pesos [10⁹ CLP]')
    ax.set_ylim(0, ax.get_ylim()[1])
    
    lines = line1 + line2 + line3 + line4 + line5
    labels = [l.get_label() for l in lines]
    ax.legend(lines, labels)
    fig.tight_layout()
    fig.show()
    fig.savefig('../pdf/total-afd-timeseries.pdf')
    return tab

def collaboration(tab1, metric='P', print_=False):
    nuniv = len(tab1)
    M = np.zeros((nuniv, nuniv))
    if not tab1.meta['computed']:
        compute_table(tab1)
    for i in range(nuniv):
        u1 = tab1[i]['University']
        tab2 = change_metric(tab1, i, metric)
        df_i = (tab2[i]['f'] - tab1[i]['f'])
        if print_:
            print(f"--- {u1} ({df_i:13.0f}) ---")
        for j in range(nuniv):
            if i == j:
                M[i, i] = df_i
                continue
            u2 = tab1[j]['University']
            tab2 = change_metric(tab1, [i, j], metric)
            df_ij = (tab2[i]['f'] - tab1[i]['f']) 
            if print_:
                print(f"{df_ij - df_i:13.0f} {df_ij:13.0f} {u2:30}")
            M[i, j] = df_ij 
    return tab1['University'], M

def plot_collaboration(u, M, size=(7.5,7.5)):
    order = np.argsort(M.diagonal())
    n = len(order)
    u = u[order]
    M = M[order,:][:,order] 
    min = np.minimum
    fig = plt.figure(1, figsize=size)
    fig.set_size_inches(*size)
    fact = np.array([M[i,:] / M[i,i] for i in range(n)])
    fig.clf()
    ax = fig.add_subplot(111, projection='3d')
    ax.view_init(azim=-41, elev=53)
    nu = np.arange(len(M))
    x, y = nu, nu
    x, y = np.meshgrid(x, y)
    x, y = x.ravel(), y.ravel()
    z = np.zeros_like(x)
    dx, dy = 0.5, 0.5
    dz = 32 * M.ravel() / 1e3
    f = fact.ravel()
    conditions = [f == 1, (f < 1) * (f > 0), f > 1, f < 0]
    colors = np.full(np.shape(dz), 'green')
    colors[f == 1] = 'blue'
    colors[f < 1] = 'y'
    colors[f < 0] = 'red'
    ax.bar3d(x, y, z, dx, dy, min(60, dz), color=colors)
    ax.set_zlim(0, 60)
    ax.set_xlim(0, 32)
    ax.set_ylim(-4, 27)
    ax.set_axis_off()
    ax.grid(False)
    ax.text(28, 13.5, 0, 'PI', zdir='y', va='center_baseline', ha='center')
    ax.text(13.5, -1, 0, 'CoI', zdir='x', va='center_baseline', ha='center')
    for i, t in zip(nu, u):
        ax.text(30, i, 0, t, zdir='x', va='center_baseline',fontsize=8)
    for i, t in zip(nu, u):
        ax.text(i, -3, 0, t, zdir='y', ha='right', va='center_baseline', fontsize=8)
    fig.tight_layout()
    fig.show()
    return fig

def science_incentives(tab, p=0.02, include_caption=False):   
    """Determine the marginal earnings for an additional paper, research
project, or post-grad professor.

Arguments:
    tab:
        AFD table
    p:
        Yearly increase of total AFD (in constant Chilean pesos).

Returns:
    None

Side effects:
    LaTex table
        
    """
    year = tab.meta['year']
    if not tab.meta['computed']:
        compute_table(tab)
    filename = '../tex/tab-incentives.tex'
    nl, tnl = "\n", "\\\\"
    mc = "\\multicolumn{2}{c}"
    caption = '\\caption'
    label = '\\label'
    with open(filename, 'w') as out:
        if include_caption:
            out.write('\\begin{table}\n')
            out.write(f"{caption}{{Additional earnings in {year} and subsequent years if a University had improved the following metrics in {year-1}: one additional full-time contract for a post-graduate professor, one additional ongoing research grant, and one additional Web of Science (ex-ISI) publication. It is assumed that the total State funding will continue to grow 2\% per year. Grants and professor contracts cumulate earnings of the same magnitude for each year there are active. Publications that involve collaborations between traditional Universities generate different marginal earnings.}}{nl}")
            out.write(f"{label}{{tab:incentives:2000}}{nl}")
        out.write('\\begin{tabular}{l rr rr rr}\n')
        out.write('\\hline\\hline\n')
        out.write(f"{'university':30} & {nl}")
        out.write(f"{mc + '{postgraduate staff}':>64} & {nl}")
        out.write(f"{mc + '{research grant}':>98} & {nl}")
        out.write(f"{mc + '{WoS publication}':>132} {tnl}{nl}")
        out.write(f"{'':30} & ")
        out.write(f"{year:13} & {'all years':15} & ")
        out.write(f"{year:13} & {'all years':15} & ")
        out.write(f"{year:13} & {'all years':15} {tnl}{nl}")
        out.write(f"{'':30} & ")
        out.write(f"{'':13} & {'[CLP]':15} & ")
        out.write(f"{'':13} & {'[CLP]':15} & ")
        out.write(f"{'':13} & {'[CLP]':15} {tnl}{nl}")
        out.write('\\hline\n')
        f = 1 / (1 - 0.95 * (1+p))
        for k in range(len(tab)):
            tab1 = change_metric(tab, k, 'Sp', increment=1) 
            df31 = 1e3 * (tab1[k]['f'] - tab[k]['f'])
            tab1 = change_metric(tab, k, 'G', increment=1) 
            df41 = 1e3 * (tab1[k]['f'] - tab[k]['f'])
            tab1 = change_metric(tab, k, 'P', increment=1) 
            df51 = 1e3 * (tab1[k]['f'] - tab[k]['f'])
            u = tab[k]['University']
            out.write(f"{u:30} & ")
            out.write(f"{df31:13,.0f} & {df31*f:15,.0f} & ")
            out.write(f"{df41:13,.0f} & {df41*f:15,.0f} & ")
            out.write(f"{df51:13,.0f} & {df51*f:15,.0f} {tnl}{nl}")
        out.write('\\hline\n')
        out.write('\\end{tabular}\n')
        if include_caption:
            out.write('\\end{table}')
    print(f'Incentives written into {filename}')

def plot_arrow(ax, x, y, /, *, label=None):
    color = ax.plot([x[0]], [y[0]], '-', label=label)[0].get_color()
    #ax.arrow(x[-2], y[-2], x[-1] - x[-2], y[-1] - y[-2], color=color,
    #    length_includes_head=True) # , head_width=0.1, overhang=0.5)
    ax.annotate('', xytext=(x[0], y[0]), xy=(x[-1], y[-1]), 
        arrowprops=dict(arrowstyle='->', color=color))

def evolution(afd_tables, name):
    universities = afd_tables[-1]['University']
    colnames = afd_tables[-1].colnames
    result = {u: 
        {n: np.array([tab[j][n] for tab in afd_tables])
            for n in colnames}
                for j, u in enumerate(universities)} 
    fig = plt.figure(1)
    fig.clf()
    ax11 = fig.add_subplot(223)
    ax11.set_ylim(0, 0.99)
    ax11.set_xlim(0, 34)
    ax11.set_xlabel('undergrads per prof')
    ax11.set_ylabel('PhD. prof fraction')

    ax12 = fig.add_subplot(224)
    ax12.set_yticks([])
    ax12.set_xlabel('publications per prof')
    ax12.set_ylim(0, 0.99)
    ax12.set_xlim(0, 1.25)

    ax21 = fig.add_subplot(221)
    ax21.set_xticks([])
    ax21.set_ylim(0, 1.25)
    ax21.set_xlim(0, 34)
    ax21.set_ylabel('publications per prof')

    for univ in universities[:-2]:
        rg = slice(None) # [-4,-3,-2,-1]
        res = result[univ]
        x1, x2, x3 = res['x2'], res['x3'] , res['x5']
        plot_arrow(ax21, x1, x3)
        plot_arrow(ax12, x3[rg], x2[rg])
        plot_arrow(ax11, x1[rg], x2[rg], label=univ)
    fig.legend(ncol=2, frameon=False, fontsize=6)
    fig.tight_layout()
    fig.subplots_adjust(wspace=0, hspace=0)
    fig.show()
    return result


if __name__ == "__main__":
    if 'init' in sys.argv:
        afd_tables = [read_table(y) for y in range(2006, 2021)]
        afd_table = afd_tables[-1]
    if 'science_incentives' in sys.argv or 'all' in sys.argv:
        science_incentives(afd_table, p=0.02)
    if 'trend' in sys.argv or 'all' in sys.argv:
        tab = variation(afd_tables)
    if 'collaboration' in sys.argv or 'all' in sys.argv:
        ax, M = collaboration(afd_table)
